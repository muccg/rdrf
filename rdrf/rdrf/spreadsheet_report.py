from rdrf.utils import get_cde_value
import openpyxl as xl
import logging
import json
from collections import OrderedDict
from rdrf.utils import evaluate_generalised_field_expression
from rdrf.dynamic_data import DynamicDataWrapper
import uuid
from django.core.servers.basehttp import FileWrapper
from rdrf.models import Registry, RegistryForm, Section, CommonDataElement
from rdrf.utils import cached
import functools

logger = logging.getLogger("registry_log")

def attempt(func):
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception, ex:
            logger.error("Longitudinal report error: function = %s args = %s kwargs = %s" % func.__name__,
                                                                                            args,
                                                                                            kwargs,
                                                                                            ex)
    return wrapper

def default_time_window():
    from datetime import datetime, timedelta
    one_year = timedelta(days=365)
    today = datetime.now()
    one_year_ago = today - one_year
    return (one_year_ago, today)


class SpreadSheetReport(object):

    def __init__(self,
                 query_model,
                 testing=False):
        self.query_model = query_model
        self.registry_model = query_model.registry
        self.projection_list = json.loads(query_model.projection)
        self.longitudinal_column_map = self._build_longitudinal_column_map()
        logger.debug("longitudinal map = %s" % self.longitudinal_column_map)
        self.output_filename = self._generate_filename()
        self.work_book = xl.Workbook()
        self.testing = testing
        self.current_sheet = None
        self.current_row = 1
        self.current_col = 1
        self.time_window = default_time_window()
        self.patient_fields = self._get_patient_fields()

    # Public interface
    def run(self):
        self._generate()
        self.work_book.save(self.output_filename)

    # Private

    def _get_patient_fields(self):
        from registry.patients.models import Patient
        patient_fields = set([ field.name for field in Patient._meta.get_fields()])
        return patient_fields


    def _generate_filename(self):
        return "/tmp/%s.xlsx" % uuid.uuid4()

    def _build_longitudinal_column_map(self):
        d = {}
        for cde_dict in self.projection_list:
            if cde_dict["longitudinal"]:
                form_name = cde_dict["formName"]
                section_code = cde_dict["sectionCode"]
                if (form_name, section_code) in d:
                    d[(form_name, section_code)].append(cde_dict["cdeCode"])
                else:
                    d[(form_name, section_code)] = [cde_dict["cdeCode"]]
        return d

    # Movement helper functions

    def _next_cell(self):
        self.current_col += 1

    def _next_row(self):
        self.current_row += 1
        self.current_col = 1

    def _reset(self):
        self.current_col = 1
        self.current_row = 1

    # Writing

    def _write_cell(self, value):
        # write value at current position
        cell = self.current_sheet.cell(
            row=self.current_row, column=self.current_col)
        cell.value = value
        self._next_cell()

    @attempt
    def _generate(self):
        config = json.loads(self.query_model.sql_query)
        static_sheets = config["static_sheets"]
        for static_sheet_dict in static_sheets:
            name = self.registry_model.code.upper() + static_sheet_dict["name"]
            columns = static_sheet_dict["columns"]
            self._create_static_sheet(name, columns)
        # these columns are always included prior
        universal_columns = config["universal_columns"]
        # to the longitudinal data
        for form_model in self.registry_model.forms:
            for section_model in form_model.section_models:
                if self._section_in_report(form_model, section_model):
                    self._create_longitudinal_section_sheet(
                        universal_columns, form_model, section_model)

    def _section_in_report(self, form_model, section_model):
        for cde_dict in self.projection_list:
            if form_model.name == cde_dict["formName"] and section_model.code == cde_dict["sectionCode"]:
                return True

    def _create_static_sheet(self, name, columns):
        self._create_sheet(name)
        self._write_header_row(columns)
        self._next_row()

        for patient in self._get_patients():
            patient_record = self._get_patient_record(patient)
            self._write_row(patient, patient_record, columns)
            self._next_row()

    def _write_row(self, patient, patient_record, columns):
        for column in columns:
            value = evaluate_generalised_field_expression(self.registry_model,
                                                          patient,
                                                          self.patient_fields,
                                                          column,
                                                          patient_record)
                                                          
            self._write_cell(value)

    def _get_timestamp_from_snapshot(self, snapshot):
        if "timestamp" in snapshot:
            return snapshot["timestamp"]
    @attempt
    @cached
    def _get_cde_value_from_snapshot(self, snapshot, form_model, section_model, cde_model):
        patient_record = snapshot["record"]
        return get_cde_value(form_model, section_model, cde_model, patient_record)

    def _create_longitudinal_section_sheet(self, universal_columns, form_model, section_model):
        sheet_name = self.registry_model.code.upper() + section_model.code
        logger.debug("write longitudinal sheet %s" % sheet_name)
        self._create_sheet(sheet_name)
        # this just writes out bit at the beginning 
        self._write_header_universal_columns(universal_columns)
        first_longitudinal_date_col = self.current_col
        self._next_row()
        max_snapshots = 0
        cde_codes = self.longitudinal_column_map.get(
            (form_model.name, section_model.code), [])

        logger.debug("longitudinal cde codes = %s" % cde_codes)
        for patient in self._get_patients():
            patient_record = self._get_patient_record(patient)
            self._write_row(patient, patient_record, universal_columns)
            num_snapshots = self._write_longitudinal_row(patient, form_model, section_model, cde_codes)
            if num_snapshots > max_snapshots:
                max_snapshots = num_snapshots
            self._next_row()

        # write longitudinal header row now we know max width
        self.current_col = first_longitudinal_date_col
        self.current_row = 1
        column_prefix = "%s/%s" % (form_model.name.upper(),
                                   section_model.code)
        logger.debug("max snapshots = %s" % max_snapshots)
        date_column_name = "DATE_%s" % column_prefix
        if max_snapshots == 0:
            self._write_cell(date_column_name)
            for cde_code in cde_codes:
                self._write_cell(cde_code)
            self._write_cell("NO DATA RECORDED!")
        
        else:
            for i in range(max_snapshots):
                self._write_cell(date_column_name)
                for cde_code in cde_codes:
                    self._write_cell(cde_code)

    def _create_sheet(self, title):
        sheet = self.work_book.create_sheet()
        sheet.title = title
        self.current_sheet = sheet
        self._reset()

    def _write_header_row(self, columns):
        for column_name in columns:
            self._write_cell(self._header_name(column_name))

    def _header_name(self, column_name):
        if column_name.startswith("@"):
            return column_name[1:]
        else:
            return column_name


    def _get_patients(self):
        from registry.patients.models import Patient
        return Patient.objects.filter(rdrf_registry__in=[self.registry_model]).order_by("id")

    @cached
    def _get_patient_record(self, patient, collection="cdes"):
        wrapper = DynamicDataWrapper(patient)
        wrapper._set_client()
        return wrapper.load_dynamic_data(self.registry_model.code, collection, flattened=False)

    def _write_header_universal_columns(self, universal_columns):
        for column_name in universal_columns:
            self._write_cell(self._header_name(column_name))

    def _write_longitudinal_row(self, patient, form_model, section_model, cde_codes):
        num_snapshots = 0
        
        for snapshot in self._get_snapshots(patient):
            num_snapshots += 1
            timestamp = self._get_timestamp_from_snapshot(snapshot)
            values = []
            for cde_code in cde_codes:
                cde_model = CommonDataElement.objects.get(code=cde_code)
                value = self._get_cde_value_from_snapshot(snapshot,
                                                          form_model,
                                                          section_model,
                                                          cde_model)
                values.append(value)
            self._write_cell(timestamp)
            for value in values:
                self._write_cell(value)
                               #                    snap1        snap2      
        return num_snapshots   #  2 snapshots = [date A B C][date A B C] - we return number so we can write the header ...

    @cached
    def _get_snapshots(self, patient):
        wrapper = DynamicDataWrapper(patient)
        wrapper._set_client()
        history_collection = wrapper._get_collection(
            self.registry_model.code, "history")
        lower_bound, upper_bound = self._get_timestamp_bounds()

        patient_snapshots = history_collection.find({"django_id": patient.pk,
                                                     "registry_code": self.registry_model.code,
                                                     "django_model": "Patient",
                                                     "record_type": "snapshot"})
        return patient_snapshots

    def _get_timestamp_bounds(self):
        dt_lower, dt_upper = self.time_window
        lower_bound = dt_lower
        upper_bound = dt_upper
        return (lower_bound, upper_bound)
