import openpyxl as xl
import logging
import json
import functools
from rdrf.helpers.utils import get_cde_value
from rdrf.helpers.utils import cached
from rdrf.db.dynamic_data import DynamicDataWrapper
from rdrf.models.definition.models import CommonDataElement, ClinicalData
from rdrf.db.generalised_field_expressions import GeneralisedFieldExpressionParser
from django.conf import settings

logger = logging.getLogger(__name__)


class Cache:
    LIMIT_SNAPSHOT = 2000
    LIMIT_CURRENT = 2000

    def __init__(self):
        self.snapshots = {}
        self.current = {}

    def _get_data(self, patient, name, cached_data, limit, retriever):
        if patient.id in cached_data:
            return cached_data[patient.id]
        else:
            if len(cached_data) == self.LIMIT_CURRENT:
                k = list(cached_data.keys())[0]
                del cached_data[k]

            patient_data = retriever(patient)
            cached_data[patient.id] = patient_data
            return patient_data

    def get_current(self, patient, current_retriever):
        return self._get_data(
            patient,
            "current",
            self.current,
            self.LIMIT_CURRENT,
            current_retriever)

    def get_snapshots(self, patient, snapshots_retriever):
        return self._get_data(
            patient,
            "snapshots",
            self.snapshots,
            self.LIMIT_SNAPSHOT,
            snapshots_retriever)


def attempt(func):
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception as ex:
            logger.error("report error with %s: %s" % (func.__name__, ex))
    return wrapper


def default_time_window():
    from datetime import datetime, timedelta
    one_year = timedelta(days=365)
    today = datetime.now()
    one_year_ago = today - one_year
    return (one_year_ago, today)


class SpreadSheetReport:
    def __init__(self, query_model, humaniser):
        self.query_model = query_model
        self.humaniser = humaniser
        self.registry_model = query_model.registry
        self.projection_list = json.loads(query_model.projection)
        self.longitudinal_column_map = self._build_longitudinal_column_map()
        self.work_book = xl.Workbook()
        self.current_sheet = None
        self.current_row = 1
        self.current_col = 1
        self.time_window = default_time_window()
        self.patient_fields = self._get_patient_fields()
        self.cde_model_map = {}
        self.cache = Cache()
        self._universal_column_map = {}

        self.gfe_func_map = {}
        self.parser = GeneralisedFieldExpressionParser(self.registry_model)

    # Public interface
    def run(self, output_filename):
        self._generate()
        self.work_book.save(output_filename)

    # Private

    def _add_cde_models(self, cde_codes):
        for cde_code in cde_codes:
            if cde_code not in self.cde_model_map:
                self.cde_model_map[
                    cde_code] = CommonDataElement.objects.get(code=cde_code)

    def _get_patient_fields(self):
        from registry.patients.models import Patient
        patient_fields = set(
            [field.name for field in Patient._meta.get_fields()])
        return patient_fields

    def _build_longitudinal_column_map(self):
        d = {}
        for cde_dict in self.projection_list:
            if cde_dict["longitudinal"]:
                form_name = cde_dict["formName"]
                section_code = cde_dict["sectionCode"]
                if (form_name, section_code) in d:
                    d[(form_name, section_code)].append(cde_dict["cdeCode"])
                else:
                    d[(form_name, section_code)] = [cde_dict["cdeCode"]]
        return d

    # Movement helper functions

    def _next_cell(self):
        self.current_col += 1

    def _next_row(self):
        self.current_row += 1
        self.current_col = 1

    def _reset(self):
        self.current_col = 1
        self.current_row = 1

    # Writing

    def _write_cell(self, value):
        # write value at current position
        cell = self.current_sheet.cell(
            row=self.current_row, column=self.current_col)
        try:
            cell.value = value
        except Exception as ex:
            logger.error("error writing value %s to cell: %s" % (value, ex))
            cell.value = "?ERROR?"
        self._next_cell()

    def _generate(self):
        config = json.loads(self.query_model.sql_query)
        static_sheets = config["static_sheets"]
        for static_sheet_dict in static_sheets:
            name = self.registry_model.code.upper() + static_sheet_dict["name"]
            columns = static_sheet_dict["columns"]
            self._create_static_sheet(name, columns)
        # these columns are always included prior
        universal_columns = config["universal_columns"]
        # to the longitudinal data
        for form_model in self.registry_model.forms:
            for section_model in form_model.section_models:
                if self._section_in_report(form_model, section_model):
                    self._create_longitudinal_section_sheet(
                        universal_columns, form_model, section_model)

    def _section_in_report(self, form_model, section_model):
        for cde_dict in self.projection_list:
            if form_model.name == cde_dict["formName"] and section_model.code == cde_dict["sectionCode"]:
                return True

    def _create_static_sheet(self, name, columns):
        self._create_sheet(name)
        self._write_header_row(columns)
        self._next_row()

        for patient in self._get_patients():
            patient_record = self.cache.get_current(patient, self._get_patient_record)
            self._write_row(patient, patient_record, columns)
            self._next_row()

    def _write_row(self, patient, patient_record, columns):
        for column in columns:
            value_retriever = self._get_value_retriever(column)
            value = value_retriever(patient, patient_record)
            self._write_cell(value)

    def _get_timestamp_from_snapshot(self, snapshot):
        if "timestamp" in snapshot:
            return snapshot["timestamp"]

    def _get_cde_value_from_snapshot(self, snapshot, form_model, section_model, cde_model):
        patient_record = snapshot["record"]
        if patient_record is None:
            return ""
        try:
            return self._human(
                form_model,
                section_model,
                cde_model,
                get_cde_value(
                    form_model,
                    section_model,
                    cde_model,
                    patient_record))
        except Exception as ex:
            patient_id = patient_record["django_id"]
            from registry.patients.models import Patient
            patient_model = Patient.objects.get(id=patient_id)
            logger.error("Error getting cde %s/%s/%s for patient %s snapshot: %s" %
                         (form_model.name, section_model.code, cde_model.code, getattr(patient_model, settings.LOG_PATIENT_FIELDNAME), ex))

            return "?ERROR?"

    @cached
    def _human(self, form_model, section_model, cde_model, raw_cde_value):
        if not isinstance(raw_cde_value, type([])):
            return self.humaniser.display_value2(
                form_model, section_model, cde_model, raw_cde_value)
        else:
            return ",".join([str(self.humaniser.display_value2(
                form_model, section_model, cde_model, x)) for x in raw_cde_value])

    def _get_value_retriever(self, column):
        if column in self.gfe_func_map:
            return self.gfe_func_map[column]
        else:
            value_retriever = self.parser.parse(column)
            self.gfe_func_map[column] = value_retriever
            return value_retriever

    def _write_universal_columns(self, patient, patient_record, universal_columns):
        patient_id = patient.pk
        if patient_id in self._universal_column_map:
            for column in universal_columns:
                self._write_cell(self._universal_column_map[patient_id][column])
        else:
            self._universal_column_map[patient_id] = {}
            for column in universal_columns:
                value_retriever = self._get_value_retriever(column)
                value = value_retriever(patient, patient_record)
                self._write_cell(value)
                self._universal_column_map[patient_id][column] = value

    def _create_longitudinal_section_sheet(self, universal_columns, form_model, section_model):
        sheet_name = self.registry_model.code.upper() + section_model.code
        sheet_name = sheet_name[:30]  # 31 max char sheet name size ...
        self._create_sheet(sheet_name)
        # this just writes out bit at the beginning
        self._write_header_universal_columns(universal_columns)
        first_longitudinal_date_col = self.current_col
        self._next_row()
        max_snapshots = 0
        cde_codes = self.longitudinal_column_map.get(
            (form_model.name, section_model.code), [])

        self._add_cde_models(cde_codes)

        for patient in self._get_patients():
            patient_record = self.cache.get_current(patient, self._get_patient_record)
            self._write_universal_columns(patient, patient_record, universal_columns)
            num_snapshots = self._write_longitudinal_row(
                patient, patient_record, form_model, section_model, cde_codes)
            if num_snapshots > max_snapshots:
                max_snapshots = num_snapshots
            self._next_row()

        # write longitudinal header row now we know max width
        self.current_col = first_longitudinal_date_col
        self.current_row = 1
        column_prefix = "%s/%s" % (form_model.name.upper(),
                                   section_model.code)
        date_column_name = "DATE_%s" % column_prefix
        if max_snapshots == 0:
            self._write_cell(date_column_name)
            for cde_code in cde_codes:
                self._write_cell(cde_code)
            self._write_cell("NO DATA RECORDED!")

        else:
            for i in range(max_snapshots):
                self._write_cell(date_column_name)
                for cde_code in cde_codes:
                    self._write_cell(cde_code)

    def _create_sheet(self, title):
        sheet = self.work_book.create_sheet()
        sheet.title = title
        self.current_sheet = sheet
        self._reset()

    def _write_header_row(self, columns):
        for column_name in columns:
            self._write_cell(self._header_name(column_name))

    def _header_name(self, column_name):
        if column_name.startswith("@"):
            return column_name[1:]
        else:
            return column_name

    def _get_patients(self):
        from registry.patients.models import Patient
        return Patient.objects.filter(rdrf_registry__in=[self.registry_model]).order_by("id")

    def _get_patient_record(self, patient, collection="cdes"):
        wrapper = DynamicDataWrapper(patient)
        return wrapper.load_dynamic_data(self.registry_model.code, collection, flattened=False)

    def _write_header_universal_columns(self, universal_columns):
        for column_name in universal_columns:
            self._write_cell(self._header_name(column_name))

    def _write_longitudinal_row(
            self,
            patient,
            patient_record,
            form_model,
            section_model,
            cde_codes):
        num_blocks = 0  # a "block" is all cdes in one snapshot or the current set of cdes
        # we will write the current set of cdes last

        # for snapshot in self._get_snapshots(patient):
        for snapshot in self.cache.get_snapshots(patient, self._get_snapshots):
            num_blocks += 1
            timestamp = self._get_timestamp_from_snapshot(snapshot)
            values = []
            for cde_code in cde_codes:
                cde_model = self.cde_model_map[cde_code]
                value = self._get_cde_value_from_snapshot(snapshot,
                                                          form_model,
                                                          section_model,
                                                          cde_model)
                values.append(value)
            self._write_cell(timestamp)
            for value in values:
                self._write_cell(value)
        # now write the current block
        current_timestamp = self._get_timestamp_from_current_form_data(
            patient_record, form_model)
        self._write_cell(current_timestamp)
        # TODO fix me The following is incomplete?
        # for cde_code in cde_codes:
        #     cde_model = self.cde_model_map[cde_code]
        #     current_value = self._human(
        #         form_model,
        #         section_model,
        #         cde_model,
        #         self._get_cde_value_from_current_data(
        #             patient_record,
        #             form_model,
        #             section_model,
        #             cde_model))

        #                    snap1        snap2      current
        # E.g 3 blocks  = [date A B C][date A B C][date A B C] - we return
        # number so we can write the header ...
        return num_blocks

    def _get_timestamp_from_current_form_data(self, patient_record, form_model):
        if patient_record is None:
            return None
        else:
            for form_dict in patient_record["forms"]:
                if form_dict["name"] == form_model.name:
                    if "timestamp" in form_dict:
                        return form_dict["timestamp"]

    def _get_cde_value_from_current_data(
            self,
            patient_record,
            form_model,
            section_model,
            cde_model):
        if patient_record is None:
            return None
        try:
            return get_cde_value(form_model, section_model, cde_model, patient_record)
        except Exception as ex:
            cde = "%s/%s/%s" % (form_model.name,
                                section_model.code, cde_model.code)
            patient_id = patient_record["django_id"]
            from registry.patients.models import Patient
            patient_model = Patient.objects.get(id=patient_id)
            logger.error("Error getting current cde %s for patient %s: %s" % (cde,
                                                                              getattr(patient_model, settings.LOG_PATIENT_FIELDNAME),
                                                                              ex))
            return "??ERROR??"

    def _get_snapshots(self, patient):
        history = ClinicalData.objects.collection(self.registry_model.code, "history")
        snapshots = history.find(patient, record_type="snapshot")

        # # fixme: date filtering was never implemented, but it could
        # # be added if the storage format of history timestamps is fixed.
        # if self.time_window:
        #     after, before = self.time_window
        #     if after is not None:
        #         snapshots = snapshots.filter(data__timestamp__gte=after.isoformat())
        #     if before is not None:
        #         snapshots = snapshots.filter(data__timestamp__lte=before.isoformat())

        return list(snapshots.data())
