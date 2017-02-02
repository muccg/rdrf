import django
django.setup()

from rdrf.models import Registry
from rdrf.models import RegistryForm
from rdrf.models import Section
from rdrf.models import CommonDataElement
from rdrf.models import RDRFContext
from rdrf.models import ConsentSection
from rdrf.models import ConsentQuestion


from rdrf.contexts_api import RDRFContextManager


from rdrf.form_view import FormView
from django.test import RequestFactory

from registry.patients.models import Patient
from registry.patients.models import PatientRelative
from registry.patients.models import PatientAddress
from registry.patients.models import AddressType
from registry.groups.models import CustomUser
from registry.groups.models import WorkingGroup

import pycountry


import openpyxl as xl
import os
import sys

EN_DASH = "–"  # This actually appears in one of the PVGs for FH (PVG-FH_GENETIC_VARIANT) - so we need to fix the spreadsheet values to use it
HYPHEN = "-"

class DataDictionary:
    FIELD_NUM_COLUMN = 1
    FORM_NAME_COLUMN = 2
    SECTION_NAME_COLUMN = 3
    CDE_NAME_COLUMN = 4
    PATIENT_ID_FIELDNUM = 1
    INDEX_FIELD_NUM = 29
    INDEX_VALUE = "index"
    RELATIVE_VALUE = "relative"
    EXTERNAL_ID_COLUMN = 1
    MY_INDEX_COLUMN = 113
    RELATIONSHIP_COLUMN = 114
    HEIGHT_COLUMN = 87
    WEIGHT_COLUMN = 88


class ImporterError(Exception):
    pass


class FieldNumNotFound(ImporterError):
    pass


class FieldType:
    CDE = 1
    DEMOGRAPHICS = 2
    PEDIGREE_FORM = 3
    CONSENT = 4


def replace_bad_value(pvg_code, s):
    if s is None:
        return s

    if pvg_code == "PVG-FH_GENETIC_VARIANT":
        if HYPHEN in s:
            return s.replace(HYPHEN, EN_DASH)

    return s





DEMOGRAPHICS_TABLE = [
    # fieldnum(column for data), English, field_expression, converter func (if
    # any
    # NB ordering is changed to allow creation of objects
    (3, "Family name", "family_name"),
    (4, "Given names", "given_names"),
    (7, "Date of birth", "date_of_birth"),
    (10, "Sex", "gender", "meteor"),
    # these fields created by _import_demographics method
    (5, "Maiden name", "maiden_name"),
    (6, "Hospital/Clinic ID", "umrn"),
    (8, "Country of birth", "country_of_birth"),
    (9, "Ethnic Origin", "ethnic_origin"),
    (11, "Home Phone", "home_phone"),
    (12, "Mobile Phone", "mobile_phone"),
    (13, "Work Phone", "work_phone"),
    (14, "Email", "email"),
    (15, "Living status", "living_status"),
    # Address handled sep
    (16, "Address", "Demographics/Address/Home/Address"),
    (17, "Suburb/Town", "Demographics/Address/Home/Suburb"),
    (18, "State", "Demographics/Address/Home/State"),
    (19, "Postcode", "Demographics/Address/Home/Postcode"),
    (20, "Country", "Demographic/Address/Home/Country"),
    (2, "Centre", "working_group"),
]


CONSENT_MAP = {
    # consent label -> fieldnum
    "Adult Consent": 21,
    "Adult consent": 21,
    "Child Consent": 22,
    "Child consent": 22,
    "Clinical Trials": 23,
    "Clinical trials": 23,
    "Information": 24,
    "FCHL": 25,
    "Hyper-Lp(a)": 26,
}


class SpreadsheetImporter(object):
    # rewriting importer
    # as we need to link relatives to index patients
    # so we do two passes
    # also FH introduced form groups
    # we need to be aware of them

    def __init__(self,
                 registry_model,
                 import_spreadsheet_filepath,
                 datadictionary_sheetname,
                 datasheet_name):
        # used for logging
        self.row = None
        self.patient = None
        self.stage = None
        self.log_prefix = "IMP"

        self.registry_model = registry_model
        self.rdrf_context_manager = RDRFContextManager(registry_model)
        self.import_spreadsheet_filepath = import_spreadsheet_filepath
        self.datadictionary_sheetname = datadictionary_sheetname
        self.datadictionary_sheet = None
        self.datasheet_name = datasheet_name
        self.data_sheet = None
        self.indexes = []
        self.relatives = []
        self.family_linkage_map = {}
        self.index_patient_map = {}
        self.field_map = {}
        self.id_map = {}   # external id -> rdrf id
        self.reverse_map = {}  # rdrf id -> external id 
        self._load_workbook()
        self._load_datasheet()
        self._load_datadictionary_sheet()
        self._build_field_map()
        self.countries = pycountry.countries
        self.ids = []
        self.mongo_db_name = self._get_mongo_db_name()


    def _get_mongo_db_name(self):
        from django.conf import settings
        return settings.MONGO_DB_PREFIX + self.registry_model.code
    
        


    def rollback_mongo(self):
        from rdrf.mongo_client import construct_mongo_client
        mongo_client = construct_mongo_client()
        db = mongo_client[self.mongo_db_name]
        cdes_collection = db["cdes"]
        
        for patient_id in self.ids:
            query = {"django_id": patient_id,
                     "django_model": "Patient"}
            cdes_collection.remove(query)
            self.log("Removed mongo data for id %s" % patient_id)

        self.log("All added patient data from run rolled back")

    def log(self, msg):
        if self.patient:
            pid = self.patient.pk
            ident = "%s/%s" % (self.external_id, pid)
        else:
            ident = ""
            
        if self.row:
            print("%s STAGE [%s] ROW [%s] PATIENT [%s]: %s" % (self.log_prefix,
                                                               self.stage,
                                                               self.row,
                                                               ident,
                                                               msg))
        else:
            print("%s STAGE [%s]: %s" % (self.log_prefix,
                                         self.stage,
                                         msg))

    def _load_workbook(self):
        if not os.path.exists(self.import_spreadsheet_filepath):
            raise ImporterError("Spreadsheet file %s does not exist" %
                                self.import_spreadsheet_filepath)

        try:
            self.workbook = xl.load_workbook(self.import_spreadsheet_filepath)

        except Exception as ex:
            raise ImporterError("Could not load Import spreadsheet: %s",
                                ex)

    def _load_datadictionary_sheet(self):
        self.datadictionary_sheet = self.workbook.get_sheet_by_name(
            self.datadictionary_sheetname)

    def _load_datasheet(self):
        self.data_sheet = self.workbook.get_sheet_by_name(self.datasheet_name)

    def _get_value(self, sheet, row, column):
        return sheet.cell(row=row,
                          column=column).value

    def _build_field_map(self):
        # map RDRF "fields" to fieldnums in the spreadsheet
        d = {}
        d["patient_id"] = DataDictionary.PATIENT_ID_FIELDNUM

        sheet = self.datadictionary_sheet
        finished = False
        row_num = 3  # fields start here

        while not finished:
            field_num = self._get_value(sheet,
                                        row_num,
                                        DataDictionary.FIELD_NUM_COLUMN)

            if not field_num:
                finished = True
            else:
                form_name = self._get_value(sheet,
                                            row_num,
                                            DataDictionary.FORM_NAME_COLUMN)

                section_name = self._get_value(sheet,
                                               row_num,
                                               DataDictionary.SECTION_NAME_COLUMN)

                cde_name = self._get_value(sheet,
                                           row_num,
                                           DataDictionary.CDE_NAME_COLUMN)

                key = (form_name, section_name, cde_name)

                d[key] = field_num
                row_num += 1

        self.field_map = d

    def _get_converter_func(self, converter_name):
        # look for methods on me like  converter_<converter_name>)
        method_name = "converter_%s" % converter_name
        if hasattr(self, method_name):
            converter_func = getattr(self, method_name)
            if callable(converter_func):
                return converter_func
        raise Exception("Bad converter: %s" % converter_name)

    # converters

    def converter_meteor(self, value):
        sex_choices = {"Male": 1,
                       "Female": 2,
                       "Indeterminate": 3}

        return sex_choices.get(value, None)

    # end of converters

    def _get_demographics_field(self, row, long_field_name):
        for tup in DEMOGRAPHICS_TABLE:
            if tup[1] == long_field_name:
                field_num = tup[0]
                if len(tup) == 4:
                    converter = tup[3]
                else:
                    converter = None
                if converter is not None:
                    value = self._get_converted_value(row,
                                                      field_num,
                                                      converter)

                else:
                    value = self._get_column(field_num, row)

                return value

        raise Exception("Unknown field %s" % long_field_name)

    def _get_field_num(self, key):
        if type(key) is tuple:
            form_model, section_model, cde_model = key
            form_name = form_model.name
            section_name = section_model.display_name
            cde_name = cde_model.name

            k = (form_name, section_name, cde_name)
            if k in self.field_map:
                return self.field_map[k]
            else:
                msg = "%s/%s/%s" % (form_name, section_name, cde_name)
                raise FieldNumNotFound(msg)
        else:
            # string key
            if key in self.field_map:
                return self.field_map[key]
            else:
                raise FieldNumNotFound(key)


    def _update_address(self, patient_model, row):
        # field expression not working?
        #(16, "Address", "Demographics/Address/Home/Address"),
        #(17, "Suburb/Town", "Demographics/Address/Home/Suburb"),
        #(18, "State", "Demographics/Address/Home/State"),
        #(19, "Postcode", "Demographics/Address/Home/Postcode"),
        #(20, "Country", "Demographic/Address/Home/Country"),

        def get_val(field_num):
            return self._get_value(self.data_sheet,
                                   row,
                                   field_num)
        address = get_val(16)
        suburb = get_val(17)
        state = get_val(18)
        postcode = get_val(19)
        country = get_val(20)

        home_address_type = AddressType.objects.get(type="Home")

        try:
            country_object = self.countries.get(name=country)
            country_code = country_object.alpha2
        except Exception as ex:
            self.log("Unknown country: [%s]" % country)
            country_code = ""

        try:
            state_code = "%s-%s" % (country_code, state)
            
            address_object = PatientAddress()
            address_object.patient = patient_model
            address_object.address = address
            address_object.suburb = suburb
            address_object.postcode = postcode
            address_object.country = country_code
            address_object.address_type = home_address_type
            address_object.state = state_code
            
            address_object.save()
        except Exception as ex:
            self.log("Error saving address: %s" % ex)

    def reset(self):
        self.stage = ""
        self.patient = None
        self.row = None

    def run(self):
        try:
            
            self.process()
            self.reset()
            self.stage = "COMPLETE"
            self.log("Import completed successfully")
        
        except Exception as ex:
            self.reset()
            self.stage = "ROLLBACK!"
            self.log("Unhandled error - rollback will now occur: %s" % ex)
            self.delete_added_patients()
                
            try:
                self.rollback_mongo()
            except Exception as rex:
                self.log("Could not rollback mongo properly: %s" % rex)
            sys.exit(1)

    def delete_added_patients(self):
        self.log("Deleting all added patients!")
        # do twice to override the default "archiving"
        Patient.objects.filter(id__in=self.ids).delete()
        Patient.objects.filter(id__in=self.ids).delete()
        
            
    def process(self):
        self.stage = "SETUP"

        self.log("Starting data import ...")
        self._load_datasheet()
        row = 2  # data starts here

        # first find which rows are index patients, which relatives
        self.log("Sorting indexes from relatives ...")
        scanned = False
        while not scanned:
            if self._check_type(row, check_index=True):
                self.indexes.append(row)
                row += 1
            elif self._check_type(row, check_index=False):
                self.relatives.append(row)
                row += 1
            else:
                scanned = True

        # now begin processing for real
        self.log("Sorting finished")
        num_indexes = len(self.indexes)
        num_relatives = len(self.relatives)
        self.log("There are %s indexes and %s relatives to import" % (num_indexes,
                                                                      num_relatives))

        self._create_indexes()

        self.stage = "RELATIVES"

        self._create_relatives()
        self.reset()
        self.stage = "COMPLETE"
        self.log("Data import finished")

    def _get_column(self, field_num, row):
        return self.data_sheet.cell(row=row,
                                    column=int(field_num)).value

    def _check_type(self, row, check_index=True):
        index_cell = self._get_value(self.data_sheet,
                                     row,
                                     DataDictionary.INDEX_FIELD_NUM)

        if check_index:
            check_value = DataDictionary.INDEX_VALUE
        else:
            check_value = DataDictionary.RELATIVE_VALUE

        value = index_cell == check_value
        return value

    def _set_working_group(self, row, patient):
        # field expression wasn't working for this
        working_group_name = self._get_demographics_field(row, "Centre")
        try:
            working_group_model = WorkingGroup.objects.get(registry=self.registry_model,
                                                           name=working_group_name)
        except WorkingGroup.DoesNotExist:
            self.log("Unknown working group: %s" % working_group_name)
            working_group_model = None

        if working_group_model:
            patient.working_groups = [working_group_model]
            patient.save()
            self.log("set working group to %s" % working_group_model)

    def _is_relative(self, row):
        value = self._get_column(
            DataDictionary.INDEX_FIELD_NUM, row) == DataDictionary.RELATIVE_VALUE
        return value

    def _create_indexes(self):
        self.stage = "CREATING INDEXES"
        self.log("creating index patients")

        for row in self.indexes:
            self.stage = "CREATING INDEXES"
            self.row = row
            self.external_id = self._get_external_id(row)
            index_patient = self._import_patient(row)
            self.patient = index_patient
            self._update_id_map(self.external_id, index_patient.pk)
            self.log("Index created")

    def _update_id_map(self, external_id, rdrf_id):
        if rdrf_id in self.reverse_map:
            raise ImportError("Dupe rdrf id: %s" % rdrf_id)
        else:
            self.reverse_map[rdrf_id] = external_id
            
        if external_id in self.id_map:
            raise ImportError("Dupe ID?- %s" % external_id)
        else:
            self.id_map[external_id] = rdrf_id

    def _create_relatives(self):
        self.stage = "CREATERELATIVES"

        for row in self.relatives:
            self.row = row
            self.external_id = self._get_external_id(row)
            patient = self._create_minimal_patient(row)
            self._update_id_map(self.external_id, patient.pk)
            self.patient = patient
            self.log("Created relative")
            self._import_patient(row, patient_to_update=patient)
            self.stage = "UPDATEIDMAP"
            self.stage = "LINKAGE"
            index_patient = self._get_index_patient(row)
            if index_patient is not None:
                self._link_relative(index_patient, patient)
            else:
                self.log("Could not link to my index as index was not found")

    def _get_index_patient(self, row):
        my_index_external_id = self._get_value(self.data_sheet,
                                               row,
                                               DataDictionary.MY_INDEX_COLUMN)

        rdrf_index_pk = self.id_map.get(my_index_external_id, None)
        if rdrf_index_pk is None:
            self.log("Could not locate MYINDEX??")
            return None
        else:
            try:
                return Patient.objects.get(pk=rdrf_index_pk)
            except Patient.DoesNotExist:
                self.log("No matching index patient for patient id = %s" %
                         rdrf_index_pk)
                return None

    def _link_relative(self, index_patient, relative_patient_model):
        self.stage = "ADDRELATIVE"
        patient_relative_model = PatientRelative()
        patient_relative_model.family_name = relative_patient_model.family_name
        patient_relative_model.given_names = relative_patient_model.given_names
        patient_relative_model.date_of_birth = relative_patient_model.date_of_birth
        patient_relative_model.sex = relative_patient_model.sex
        patient_relative_model.relationship = self._get_relationship_to_index()
        patient_relative_model.living_status = relative_patient_model.living_status
        # set index of this relative
        patient_relative_model.patient = index_patient

        # as all imported relatives are also patients in the registry
        # point this PatientRelative to the relative
        patient_relative_model.relative_patient = relative_patient_model

        patient_relative_model.save()
        if relative_patient_model is not None and index_patient is not None:
            self.log("Linked relative %s to index %s" % (relative_patient_model.pk,
                                                         index_patient.pk))

    def _get_relationship_to_index(self):
        relationship = self._get_value(self.data_sheet,
                                       self.row,
                                       DataDictionary.RELATIONSHIP_COLUMN)

        return relationship

    def _get_field(self, row, field):
        if type(field) is type(basestring):
            # demographics
            field_num = self._get_field_num(field)
            value = self._get_value(self.data_sheet,
                                    row,
                                    field_num)
            return value

    def _import_demographics_data(self, patient, row):
        updates = []
        self.stage = "WORKINGGROUP"
        self._set_working_group(row, patient)
        self.stage = "DEMOGRAPHICS"
        for t in DEMOGRAPHICS_TABLE:
            has_converter = len(t) == 4
            field_num = t[0]
            field_expression = t[2]
            value = self._get_column(field_num, row)
            if has_converter:
                converter_name = t[3]
                converter_func = self._get_converter_func(converter_name)
                value = converter_func(value)
            updates.append((field_expression, value))

        context_model = self.rdrf_context_manager.get_or_create_default_context(
            patient, new_patient=True)

        self.log("context for patient = %s" % context_model.pk)
        patient.update_field_expressions(
            self.registry_model, updates, context_model=context_model)

        self.stage = "ADDRESS"
        self._update_address(patient, row)
        self.stage = "DEMOGRAPHICS"

    def _import_pedigree_data(self, patient, row):
        self.patient = patient
        self.row = row
        self.stage = "PEDIGREE"

    def _get_context_for_patient(self, patient_model):
        if not self.registry_model.has_feature("contexts"):
            return patient_model.default_context

        for context_model in patient_model.context_models:
            if context_model.context_form_group and context_model.context_form_group.context_type == "F":
                return context_model

    def _get_converted_value(self, row, field_num,  converter=None):
        value = self._get_column(field_num, row)
        if converter:
            converter_func = self._get_converter_func(converter)
            value = converter_func(value)
        return value

    def _create_minimal_patient(self, row):
        self.stage = "MINIMAL"
        patient = Patient()
        self.row = row
        self.patient = patient
        family_name = self._get_demographics_field(row, "Family name")
        given_names = self._get_demographics_field(row, "Given names")
        date_of_birth = self._get_demographics_field(row, "Date of birth")
        sex = self._get_demographics_field(row, "Sex")
        patient.family_name = family_name
        patient.given_names = given_names
        patient.date_of_birth = date_of_birth
        patient.sex = sex
        patient.consent = True  # to satisfy validation ...
        patient.save()
        self.ids.append(patient.pk)
        
        patient.rdrf_registry = [self.registry_model]
        patient.save()

        return patient

    def _import_patient(self, row, patient_to_update=None):
        self.stage = "MINIMAL"
        self.row = row
        if patient_to_update is None:
            patient = self._create_minimal_patient(row)
            self.patient = patient
        else:
            patient = patient_to_update
            
        self.stage = "DEMOGRAPHICS"
        self._import_demographics_data(patient, row)
        self._import_consents(patient, row)

        self.stage = "CLINICAL"
        self._import_clinical_data(patient, row)

        return patient

    def _import_clinical_data(self, patient, row):
        for context_model, form_model in self._get_forms_and_contexts(patient):
            self._import_clinical_form(form_model, patient, context_model, row)

    def _get_forms_and_contexts(self, patient_model):
        # get forms in fixed groups
        results = []
        for context_form_group in self.registry_model.context_form_groups.all():
            if context_form_group.is_default and context_form_group.context_type == "F":
                try:
                    context_model = RDRFContext.objects.get(object_id=patient_model.pk,
                                                            registry=self.registry_model,
                                                            context_form_group=context_form_group)
                except RDRFContext.DoesNotExist:
                    # should not happen as contexts for fixed groups created
                    # when patient craeted
                    raise Exception("could not find context for patient %s cfg %s" % (
                        patient_model, context_form_group))

                except RDRFContext.MultipleObjectsReturned:
                    raise Exception("too many contexts for patient %s cfg %s" %
                                    (patient_model, context_form_group))

                for form_model in context_form_group.form_models:
                    results.append((context_model, form_model))
        return results

    def _import_consents(self, patient, row):
        self.stage = "CONSENTS"
        self.patient = patient
        self.row = row
        # NB these are FH specific
        self._import_consent("Registry Consent", "Adult consent")
        self._import_consent("Registry Consent", "Child consent")
        self._import_consent("Optional Consents", "Clinical trials")
        self._import_consent("Optional Consents", "Information")

        self._import_consent("Registry Subset", "FCHL")
        self._import_consent("Registry Subset", "Hyper-Lp(a)")

    def _import_consent(self, section_label, question_label):
        self.stage = "CONSENT/%s" % section_label
        field_num = CONSENT_MAP.get(question_label, None)
        if field_num is None:
            raise Exception("Unknown consent question: %s" % question_label)

        answer = self._get_value(self.data_sheet,
                                 self.row,
                                 field_num)

        if answer is None:
            answer = "False"

        if answer.lower() in ["true", "yes", "y"]:
            answer = True
        else:
            answer = False

        self._import_consent_question(section_label, question_label, answer)

    def _import_consent_question(self, section_label, question_label, answer):
        consent_section_model = self._get_consent_section(section_label)
        consent_question_model = ConsentQuestion.objects.get(section=consent_section_model,
                                                             question_label=question_label)

        answer_field_expression = "Consents/%s/%s/answer" % (consent_section_model.code,
                                                             consent_question_model.code)

        self._eval(answer_field_expression, answer)

    def _get_consent_section(self, section_label):
        try:
            consent_section_model = ConsentSection.objects.get(registry=self.registry_model,
                                                               section_label=section_label)
            return consent_section_model
        except ConsentSection.DoesNotExist:
            self.log("ConsentSection with label %s does not exist" %
                     section_label)
            return None

    def _eval(self, field_expression, value):
        self.patient.evaluate_field_expression(self.registry_model,
                                               field_expression,
                                               value=value)

    def _convert_cde_value(self, cde_model, spreadsheet_value):
        if cde_model.pv_group:
            spreadsheet_value = replace_bad_value(cde_model.pv_group.code, spreadsheet_value)
            for pv in cde_model.pv_group.permitted_value_set.all():
                if spreadsheet_value == pv.value:
                    return pv.code
            return None
        else:
            return spreadsheet_value

    def _import_clinical_form(self, form_model, patient_model, context_model, row):
        self.stage = "CLINICAL"
        self.patient = patient_model
        self.row = row
        field_updates = []
        for section_model in form_model.section_models:
            if not section_model.allow_multiple:
                # 1st data import sheet does not contain any multisections as
                # is 1 row per patient
                for cde_model in section_model.cde_models:
                    if self._cde_included(form_model, section_model, cde_model):
                        field_num = self._get_field_num(
                            (form_model, section_model, cde_model))
                        spreadsheet_value = self._get_value(self.data_sheet,
                                                            row,
                                                            field_num)

                        rdrf_value = self._convert_cde_value(
                            cde_model, spreadsheet_value)

                        field_expression = self._get_field_expression(
                            form_model, section_model, cde_model)

                        if "CDEBMI" in field_expression:
                            # special case
                            self._update_bmi(row,
                                             field_expression)
                            
                        else:
                            field_updates.append((field_expression, rdrf_value))

                    else:
                        self.log("Form %s Section %s CDE %s NOT INCLUDED IN SPREADSHEET" % (form_model.name,
                                                                                            section_model.display_name,
                                                                                            cde_model.name))
        if len(field_updates) > 0:
            patient_model.update_field_expressions(
                self.registry_model,
                field_updates,
                context_model
            )

    def _update_bmi(self, row, bmi_field_expression):
        self.stage = "BMI"
        height = self._get_value(self.data_sheet,
                                 row,
                                 DataDictionary.HEIGHT_COLUMN)
        weight = self._get_value(self.data_sheet,
                                 row,
                                 DataDictionary.WEIGHT_COLUMN)

        try:
            h = float(height)
            w = float(weight)
            bmi = w / (h * h)
            self._eval(bmi_field_expression, bmi)
        except Exception as ex:
            self.log("Could not update bmi: %s" % ex)
        self.stage = "CLINICAL"

    def _get_external_id(self, row):
        return self._get_column(DataDictionary.EXTERNAL_ID_COLUMN, row)

    def _cde_included(self, form_model, section_model, cde_model):
        name_tuple = (form_model.name,
                      section_model.display_name, cde_model.name)
        return name_tuple in self.field_map

    def _get_field_expression(self, form_model, section_model, cde_model):
        expression = "%s/%s/%s" % (form_model.name,
                                   section_model.code,
                                   cde_model.code)

        return expression


if __name__ == "__main__":
    registry_code = sys.argv[1]
    spreadsheet_file = sys.argv[2]
    
    registry_model = Registry.objects.get(code=registry_code)
    spreadsheet_importer = SpreadsheetImporter(registry_model,
                                               spreadsheet_file,
                                               "fh_data_dictionary_v3_fhwa.xlsx",
                                               "Sheet1")
    spreadsheet_importer.run()
