import django
django.setup()

import sys
import os
import re
from datetime import datetime

from django.db import transaction
from django.core.exceptions import ValidationError
from rdrf.models import Registry
from registry.groups.models import WorkingGroup
from registry.patients.models import Patient
from registry.patients.models import PatientAddress, AddressType
from openpyxl import load_workbook

import random
# All addresses must have a state / region in rdrf

nz_regions = {1:'otago',
              2:'taranaki',
              3:'auckland',
              4:'southland',
              5:'nelson city',
              6:'bay of plenty',
              7:'south island',
              8:'chatham islands territory',
              9:'waikato',
              10:'gisborne district',
              11:'northland',
              12:'north island',
              13:'canterbury',
              14:'manawatu-wanganui',
              15: 'tasman district',
              16: 'marlborough district',
              17: 'wellington',
              18: 'west coast',
              19: "hawke's bay"
}

# We have cities/towns but not regions in the import...
nz_city_map = {
    "Ashburton": 13,
    "Auckland": 3,
    "Blenheim": 16,
    "Christchurch": 13,
    "Dannevirke": 14,
    "Fielding": 14,
    "Foxton": 14,
    "Gisborne": 10 ,
    "Gisbourne": 10,
    "Greytown": 12,
    "Hamilton": 9,
    "Hastings": 19,
    "Hokitika": 18,
    "Kaiapoi": 13,
    "Lyttleton": 13,
    "Mangakino": 9,
    "Masterton" : 17,
    "Matamata" : 9,
    "Mosgiel": 1,
    "Motueka":15,
    "Nelson" : 15,
    "New Plymouth": 2,
    "Otaki" : 12,
    "Palmerston North": 14,
    "Papakura": 3,
    "Parakai": 3,
    "Porirua": 17,
    "Pukekohe": 3,
    "Rangiora": 13,
    "Rotorua": 6,
    "Stratford": 2,
    "Tauranga": 6,
    "Tawa": 17,
    "Timaru": 13,
    "Tokoroa": 9,
    "Wainuiomata": 17,
    "Waipawa": 19,
    "Wanganui": 14,
    "Wellington": 17,
    "Westport": 18,
    "Whakatane": 6,
    "Whangaparaoa": 3,
    "Whangarei": 11,
}

def is_rd(s):
    pattern = re.compile(r"^RD\s*\d+$")
    return pattern.match(s)

    
def get_region(place):
    x = place.strip().lower()
    for city in nz_city_map:
        if x == city.lower():
            region_num = nz_city_map[city]
            region = nz_regions[region_num]
            return region
        

class Columns:
    GIVEN_NAMES = 2
    FAMILY_NAME = 3
    DOB = 4
    SEX = 8
    ADDRESS1 = 10
    ADDRESS2 = 11  # rural address type or house name??
    ADDRESS3 = 12 # suburb
    ADDRESS4 = 13 # town
    SUBURB = 12
    TOWN = 13 # actually it's complicated - some state info mixed in
    POSTCODE = 14


class ProcessingError(Exception):

    def __init__(self, row_number, column_label, patient_model, message):
        super(ProcessingError, self).__init__(message)
        self.row_number = row_number
        self.column_label = column_label
        self.patient_model = patient_model

    def __str__(self):
        return "Error row %s column %s patient %s: %s" % (self.row_number,
                                                          self.column_label,
                                                          self.patient_model,
                                                          self.message)


class Dm1Importer(object):
    MINIMAL_FIELDS = [Columns.GIVEN_NAMES,
                      Columns.FAMILY_NAME, Columns.DOB, Columns.SEX]
    COLS = {
        # from spreadsheet sample
        # NHI	First Name	Family Name	DOB	Date of consent	Genetic Test recd.	Diagnosis	Sex	Ethnicity	Postal address 1
        # Address 2	Address 3	Address 4	Pcode	Email	Home Ph	Mobile	Participants representative
        # Relationship to participant	Address	Email	Ph	<no value> Key

        # column index --> (column label, generalised field expression/set_function name, convert function name (if any))
        # if second of pair is None - we have a special case -
        1: ("NHI", "umrn"),
        2: ("First Name", "given_names"),
        3: ("Family Name", "family_name"),
        4: ("DOB", "date_of_birth"),
        5: ("Date of consent", "set_consent"),
        6: ("Genetic Test recd.", "ClinicalData/DM1GeneticTestDetails/GeneticTestResultsAvailable", "convert_genetic_test_received"),
        7: ("Diagnosis", "ClinicalData/DM1Status/DM1Condition", "convert_dm1condition"),
        8: ("Sex", "sex"),
        9: ("Ethnicity", "set_ethnic_origin"),
        # Address spans multiple columns - special case
        10: ("Postal Address 1", None),
        11: ("Address 2", None),
        12: ("Address 3", None),
        13: ("Address 4", None),
        14: ("Pcode", None),
        15: ("Email", "email"),
        16: ("Home Ph", "home_phone"),
        17: ("Mobile", "mobile_phone"),
        # Special case also
        18: ("Participants Representative", None),
        19: ("Relationship to Participant", None),
        20: ("Address", None),
        21: ("Email", None),
        22: ("Ph", None),
        23: (None, None),
        24: ("Key", None)

    }

    def __init__(self, registry_model, excel_filename):
        self.excel_filename = excel_filename
        self.workbook = load_workbook(self.excel_filename)
        self.sheet = self.workbook.worksheets[0]

        self.registry_model = registry_model
        self.working_group = WorkingGroup.objects.get(registry=self.registry_model,
                                                      name="New Zealand")

        self.errors = []
        self.num_patients_created = 0
        self.current_row = 2  # 1 based
        self.current_patient = None
        self.log_name = "IMPORTER "  # so we can grep
        self.log_prefix = "INIT"
        self.POSTAL_ADDRESS_TYPE, created = AddressType.objects.get_or_create(type="Postal",
                                                                              description="Postal")
        if created:
            self.POSTAL_ADDRESS_TYPE.save()

        self.STATE_MAP = self._build_state_map()
        self.STATE_NAMES = self.STATE_MAP.keys()
        self.address_errors = 0

    def sep(self):
        self.log("***************************************************")

    def _build_state_map(self):
        d = {}
        import pycountry
        states = sorted(pycountry.subdivisions.get(country_code="NZ"),
                        key=lambda x: x.name)

        for state in states:
            d[state.name.lower()] = state.code
        return d

    def log(self, msg):
        print "%s %s: %s" % (self.log_name, self.log_prefix, msg)

    def _get_cell(self, current_row, column_index):
        value = self.sheet.cell(row=current_row, column=column_index).value
        return value

    def run(self):
        while self.current_row is not None:
            row = {}
            for column_index in range(1, 24):
                row[column_index] = self._get_cell(
                    self.current_row, column_index)

            if self._is_blank_row(row):
                self.log_prefix = "Finished"
                self._display_summary()
                self.current_row = None
            else:
                self.sep()
                self._process_row(row)
                self.sep()
                self.current_row += 1

    def _is_blank_row(self, row_dict):
        return all(map(lambda index: row_dict[index] is None, self.MINIMAL_FIELDS))

    def _display_summary(self):
        self.log("%s patients imported" % self.num_patients_created)
        self.log("Number of errors: %s" % len(self.errors))
        self.log("Number of addresses NOT imported: %s" % self.address_errors)
        for error_dict in self.errors:
            self._print_error(error_dict)

    def _print_error(self, error):
        self.log("ERROR: %s" % error)

    def _process_row(self, row_dict):
        self.log_prefix = "Row %s" % self.current_row
        self.log("starting to process")
        self.current_patient = self._create_minimal_patient(row_dict)
        self.log("created minimal patient %s" % self.current_patient)
        for i in self.COLS:
            if i not in self.MINIMAL_FIELDS and self.COLS[i][1] is not None:
                column_name = self.COLS[i][0]
                try:
                    self._update_field(i, row_dict)
                except Exception, ex:
                    message = "%s" % ex
                    self.errors.append(ProcessingError(self.current_row,
                                                       column_name,
                                                       self.current_patient,
                                                       message))

        self._set_address_fields(row_dict)
        self._set_parent_guardian_fields(row_dict)

    def _set_address_fields(self, row_dict):
        address1 = row_dict[Columns.ADDRESS1]

        # not unit has values like RD1 RD3 + sometimes a house name
        address2 = row_dict[Columns.ADDRESS2]
        # Must be
        # https://www.nzpost.co.nz/personal/receiving-mail/rural-delivery#types

        suburb = row_dict[Columns.SUBURB]
        town = row_dict[Columns.TOWN]
        postcode = row_dict[Columns.POSTCODE]
        state = get_region(town)

        s = state.strip().lower()
        state_code = self.STATE_MAP.get(s, None)
        if state_code is None:
            self.log("State %s is not one of: %s" % (s,
                                                     self.STATE_NAMES))

        if town and not suburb:
            suburb = town
        elif suburb and not town:
            pass
        elif town and suburb:
            self.log("town and suburb provided - using both with a ,")
            suburb = suburb + ", " + town

        else:
            self.log("no suburb / town info provided")
            

        if address2:
            address = address1 + ", " + address2
        else:
            address = address1

        self._create_address(address, suburb, state_code, postcode)

    def _create_address(self, address, suburb, state_code, postcode):
        # Demographics/Address/Postal/Address
        # Demographics/Address/Postal/Suburb
        # Demographics/Address/Postal/State
        # Demographics/Address/Posta<select class="form-control"
        # id="id_patient_address-0-state" name="patient_address-0-state">

        if not all([address, suburb, postcode, state_code]):
            self.address_errors += 1
            self.log("address info incomplete: address=[%s] suburb=[%s] postcode=[%s] state_code=[%s]" % (address,
                                                                                                          suburb,
                                                                                                          postcode,
                                                                                                          state_code))
            return

        pa = PatientAddress()
        pa.address_type = self.POSTAL_ADDRESS_TYPE
        pa.address = address
        pa.suburb = suburb
        pa.postcode = postcode
        pa.state = state_code
        pa.country = "NZ"
        pa.patient = self.current_patient
        # if save throws db exception here, we can't do subsequent queries until
        # after atomic block
        pa.save()
        self.log("created address OK")

    def set_ethnic_origin(self, raw_value):
        mapping = {"nz european": "new zealand european",
                   "nz european/maori": "nz european / maori"}

        cde_values = ["New Zealand European", "Australian", "Other Caucasian/European",
                      "Aboriginal", "Person from the Torres Strait", "Maori",
                      "NZ European / Maori", "Samoan", "Cook Islands Maori", "Tongan",
                      "Niuean", "Tokelauan", "Fijian", "Other Pacific Peoples", "Southeast Asian",
                      "Chinese", "Indian", "Other Asian", "Middle Eastern", "Latin American",
                      "Black African/African American", "Other Ethnicity", "Decline to Answer"]

        v = raw_value.lower().strip()
        if not v:
            self.current_patient.ethnic_origin = None
            self.current_patient.save()
            return

        v = mapping.get(v, v)

        other = "Other Ethnicity"
        x = other
        for eo, _ in Patient.ETHNIC_ORIGIN:
            if eo.lower() == v:
                x = eo
                break
        self.current_patient.ethnic_origin = x
        self.current_patient.save()

        # mirror in cde DM1EthnicOrigins
        pvg_value = "Other Ethnicity"
        if v not in [cde.lower() for cde in cde_values]:
            self.log(
                "DM1EthnicOrigin supplied value %s will be mapped to Other Ethnicity" % raw_value)
        for value in cde_values:
            if v == value.lower():
                pvg_value = value

        self.execute_field_expression(
            "ClinicalData/DM1EthnicOrigins/DM1EthnicOrigins", pvg_value)

    def _set_parent_guardian_fields(self, row_dict):
        self.log("set parent guardian fields - to do")

    def _update_field(self, column_index, row_dict):
        column_info = self.COLS[column_index]
        converter = None
        if len(column_info) == 2:
            column_label, field_expression = column_info
        else:
            column_label, field_expression, converter_name = column_info
            converter = getattr(self, converter_name)

        if field_expression.startswith("set_"):
            # custom set_ function
            self.log("running function %s" % field_expression)
            func = getattr(self, field_expression)
            value = row_dict[column_index]
            self.log("raw_value = %s" % value)
            if converter is not None:
                self.log("raw_value will be converted")

                self.log("value before conversion = %s" % value)
                converted = converter(value)
                self.log("value after conversion = %s" % converted)
                func(converted)
            else:
                if value is None:
                    self.log("value is None - won't set")
                else:
                    func(value)

        else:
            # generalised field expression
            value = row_dict[column_index]
            if converter is not None:
                value = converter(value)

            self.execute_field_expression(field_expression, value)

    def execute_field_expression(self, field_expression, value):
        self.current_patient.evaluate_field_expression(self.registry_model,
                                                       field_expression,
                                                       value=value)
        self.log("%s --> %s" % (field_expression, value))

    def set_consent(self, consent_date):
        if not consent_date:
            return

        # check consents: dm1consentsec01 (c2) and dm1consentsec02 (c1 and c2)
        self.execute_field_expression(
            "Consents/dm1consentsec01/c2/answer", True)
        self.execute_field_expression(
            "Consents/dm1consentsec01/c2/first_save", consent_date)
        self.execute_field_expression(
            "Consents/dm1consentsec01/c2/last_update", consent_date)

        self.execute_field_expression(
            "Consents/dm1consentsec02/c1/answer", True)
        self.execute_field_expression(
            "Consents/dm1consentsec02/c1/first_save", consent_date)
        self.execute_field_expression(
            "Consents/dm1consentsec02/c1/last_update", consent_date)

        self.execute_field_expression(
            "Consents/dm1consentsec02/c2/answer", True)
        self.execute_field_expression(
            "Consents/dm1consentsec02/c2/first_save", consent_date)
        self.execute_field_expression(
            "Consents/dm1consentsec02/c2/last_update", consent_date)

    def _get_field_value(self, our_field_name, row_dict):
        value = None
        for index in self.COLS:
            column_info = self.COLS[index]
            if len(column_info) == 2:
                column_label, our_expression = self.COLS[index]
            else:
                column_label, our_expression, _ = self.COLS[index]

            if our_field_name == our_expression:
                value = row_dict[index]
                break
        return value

    def convert_dm1condition(self, value):
        # NB the RDRF PVG range values
        v = value.lower().strip()
        mapping = {"myotonic dystrophy": "DM1ConditionDM1",
                   "proximal myotonic myopathy promm": "DM1ConditionDM2"}

        return mapping[v]

    def _create_minimal_patient(self, row_dict):
        self.log("creating patient ...")
        p = Patient()
        p.given_names = self._get_field_value("given_names", row_dict)
        p.family_name = self._get_field_value("family_name", row_dict)
        p.sex = self._convert_sex(self._get_field_value("sex", row_dict))
        p.date_of_birth = self._convert_date_of_birth(self._get_field_value("date_of_birth",
                                                                            row_dict))

        p.consent = True
        p.active = True

        p.save()
        self.log("saved minimal fields for %s" % p)

        p.rdrf_registry = [self.registry_model]
        p.working_groups = [self.working_group]
        p.save()
        self.log("assigned to %s and %s" % (self.registry_model,
                                            self.working_group))

        self.num_patients_created += 1

        return p

    def _convert_date_of_birth(self, dob):
        return dob

    def _convert_sex(self, raw_sex):
        mapping = {
            "M": 1,
            "F": 2
        }

        return mapping.get(raw_sex, None)

    def convert_genetic_test_received(self, raw_value):
        mapping = {
            "yes": "YesNoUnknownYes",
            "no":  "YesNoUnknownNo",
            # !! - this was requested in RDR-1333 ...
            "pending": "YesNoUnknownUnknown",
            "family member +ve test": "YesNoUnknownUnknown"  # ditto

        }

        if not raw_value:
            return None

        return mapping.get(raw_value.lower().strip(), None)


if __name__ == '__main__':
    excel_filename = sys.argv[1]
    if not os.path.exists(excel_filename):
        print "Excel file does not exist"
        sys.exit(1)

    registry_code = "DM1"
    try:
        registry_model = Registry.objects.get(code=registry_code)
    except Registry.DoesNotExist:
        print "DM1 registry does not exist on the site - nothing imported"
        sys.exit(1)

    try:
        with transaction.atomic():
            dm1_importer = Dm1Importer(registry_model, excel_filename)
            dm1_importer.run()
            if len(dm1_importer.errors) > 0:
                for error in dm1_importer.errors:
                    print error
                raise Exception("processing errors occurred")

            print "%s (%s percent) addresses couldn't be created" % (dm1_importer.address_errors,
                                                                     100.0 * (float(dm1_importer.address_errors) / float(dm1_importer.num_patients_created)))

            print "run successful - NO ROLLBACK!"

    except Exception, ex:
        print "Error running import (will rollback): %s" % ex
        sys.exit(1)
